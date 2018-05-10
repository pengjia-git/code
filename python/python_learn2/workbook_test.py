from openpyxl import Workbook

wb = Workbook()
ws=wb.active
ws1=wb.create_sheet("sheet_jia");
ws1.title="new title"
#print wb.sheetnames
c=ws["A1"]
